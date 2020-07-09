# python 3.7
# launch runEXE.py launch switchExp.exe
from openpyxl import Workbook, load_workbook
import subprocess
import sys
import os

dataLocation_1 = sys.argv[1]  # ex: random_graph
dataLocation_2 = sys.argv[2]  # ex: plan_graph
dataLocation_3 = sys.argv[3]  # ex: group_1
dataLocation_4 = sys.argv[4]  # ex: dataset_1_1_1
sampples_bit = sys.argv[5]  # ex: 1023(first ten graphs)
formulation = sys.argv[6]  # 1\2\3\4
callback_option = sys.argv[7]  # 0\1\2\3
relax_option = sys.argv[8]  # 0\1
ns_sep_opt = sys.argv[9]  # 0\1
LB_MaxRestart = sys.argv[10]
LB_MaxIter = sys.argv[11]
Rmin = sys.argv[12]
Rmax = sys.argv[13]
BCSolNum = sys.argv[14]
BCTime = sys.argv[15]
MIPDisplayLevel = sys.argv[16]
time_limit = sys.argv[17]  # ex: 3600
max_cut_number_lazy = sys.argv[18]  # ex: 5
epsilon_lazy = sys.argv[19]  # ex: 0.2
max_cut_number_user = sys.argv[20]  # ex: 5
epsilon_user = sys.argv[21]  # ex: 0.2
UseLocalBranch = sys.argv[22]
LB_CP_Option = sys.argv[23]
lazy_sep_opt = sys.argv[24]

os.chdir('../..')  # to ...SMP/
# print(os.getcwd())
cwd = os.getcwd()
exeAbsltLocation = cwd + '\\x64\\Release\\SMP_1271_e_c.exe'
dataAbsltLocation = cwd + '\\test\\data\\'

# D:~/SMP/test/data/random_graph/plan_random/group_1/dataset1_1_1_2/
dataAbsltLocation = dataAbsltLocation + dataLocation_1 + '\\' + \
    dataLocation_2 + '\\' + dataLocation_3 + '\\' + dataLocation_4 + '\\'

i = 1
test_graph_list = ['begin']
for b in reversed(bin(int(sampples_bit))):
    if b == 'b':
        i -= 2
        break
    elif b == str(1):
        tempDataLocation = ''
        # D:/GitHub/Repo/SMPtest/data/random_graph/plan_random/group_1/dataset1_1_1_2/animal_10_2_5_84%_
        tempDataLocation = dataAbsltLocation + 'animal_' + str(i) + '.txt'
        print('\n++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\
            ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n')
        print('graph_'+str(i) + ' START')
        subprocess.Popen([exeAbsltLocation, tempDataLocation, formulation, callback_option, relax_option,
                          ns_sep_opt, LB_MaxRestart, LB_MaxIter, Rmin, Rmax, BCSolNum, BCTime, MIPDisplayLevel,
                          time_limit, max_cut_number_lazy, epsilon_lazy, max_cut_number_user, epsilon_user,
                          UseLocalBranch, LB_CP_Option, lazy_sep_opt]).wait()
        print('graph_'+str(i)+' DONE')
        test_graph_list.append(str(i))
    i += 1

##############################
#### Begin to analyse data ###
##############################


if relax_option == 1:

    if formulation == '1':
        result_stream = open(dataAbsltLocation + "1_SCF.txt", "r")
    elif formulation == '2':
        result_stream = open(dataAbsltLocation + "1_MCF.txt", "r")
    elif formulation == '3':
        result_stream = open(dataAbsltLocation + "1_STEINER.txt", "r")
    elif formulation == '4':
        result_stream = open(dataAbsltLocation + "1_NS.txt", "r")

    wb = load_workbook(cwd+'\\test\\data\\allDatas.xlsx')
    ws = wb['testData']

    # Find the first empty row
    first_empty_row = ws.max_row + 1
    print('first_empty_row = ', first_empty_row)

    print(test_graph_list)
    test_graph_index = 1
    for line in result_stream:
        print(line)
        ws['A'+str(first_empty_row)] = dataLocation_1
        ws['B'+str(first_empty_row)] = dataLocation_2
        ws['C'+str(first_empty_row)] = dataLocation_3
        ws['D'+str(first_empty_row)] = dataLocation_4
        print(test_graph_list[test_graph_index])
        ws['E'+str(first_empty_row)] = 'animal_' + \
                   str(test_graph_list[test_graph_index])
        split_result_line = line.split()
        ws['F'+str(first_empty_row)] = split_result_line[3]
        test_graph_index += 1
        first_empty_row += 1

elif relax_option == 0:

    if formulation == '1':
        result_stream = open(dataAbsltLocation + "1_SCF.txt", "r")
    elif formulation == '2':
        result_stream = open(dataAbsltLocation + "1_MCF.txt", "r")
    elif formulation == '3':
        result_stream = open(dataAbsltLocation + "1_STEINER.txt", "r")
    elif formulation == '4':
        result_stream = open(dataAbsltLocation + "1_NS.txt", "r"

    wb=load_workbook(cwd+'\\test\\data\\allDatas.xlsx')
    ws=wb['testData']

    print(test_graph_list)
    test_graph_index=1
    for line in result_stream:
        # Find the corresponding row
        for a in ws['A']:
            this_row=a.row
            if a.value == dataLocation_1:
                if ws['B'+str(temp_row)].value == dataLocation_2:
                    if ws['C'+str(temp_row)].value == dataLocation_3:
                        if ws['D'+str(temp_row)].value == dataLocation_4:
                            if ws['E'+str(temp_row)].value == 'animal'+str(test_graph_list[test_graph_index]):
                                break
        for a in ws['1']

        first_empty_column=ws.max_column + 1
        print(line)
        ws['A'+str(first_empty_row)]=dataLocation_1
        ws['B'+str(first_empty_row)]=dataLocation_2
        ws['C'+str(first_empty_row)]=dataLocation_3
        ws['D'+str(first_empty_row)]=dataLocation_4
        print(test_graph_list[test_graph_index])
        ws['E'+str(first_empty_row)]='animal_' + \
                   str(test_graph_list[test_graph_index])
        split_result_line=line.split()
        ws['F'+str(first_empty_row)]=split_result_line[3]
        test_graph_index += 1
        first_empty_row += 1

wb.save(cwd+'\\test\\data\\allDatas.xlsx')
